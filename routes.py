import os
import sys
from flask import render_template, request, redirect, url_for, flash, session, jsonify, send_file, make_response
from app import app, db
from models import Classroom, Schedule, Incident
from datetime import datetime, timedelta
try:
    import pytz
    PYTZ_AVAILABLE = True
except ImportError:
    PYTZ_AVAILABLE = False
    # Fallback timezone handling
    from datetime import timezone
    pytz = None
import io
from urllib.parse import urljoin
from werkzeug.utils import secure_filename
import uuid

# Import optional dependencies with error handling
try:
    from pdf_generator import generate_classroom_pdf, generate_general_report, generate_availability_report
    PDF_AVAILABLE = True
except ImportError as e:
    import logging
    logging.warning(f"PDF generation not available: {e}")
    generate_classroom_pdf = generate_general_report = generate_availability_report = None
    PDF_AVAILABLE = False

try:
    from qr_generator import generate_qr_code
except ImportError as e:
    import logging
    logging.warning(f"QR code generation not available: {e}")
    generate_qr_code = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError as e:
    import logging
    logging.warning(f"Excel functionality not available: {e}")
    openpyxl = Font = Alignment = PatternFill = None
    EXCEL_AVAILABLE = False

ADMIN_PASSWORD = "senai103103"
# All files are now stored in PostgreSQL database, no local file storage
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_excel_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS

def is_admin_authenticated():
    return session.get('admin_authenticated', False)

def require_admin_auth(f):
    def decorated_function(*args, **kwargs):
        if not is_admin_authenticated():
            flash('Acesso negado. Autenticação necessária.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    classrooms = Classroom.query.all()
    return render_template('index.html', classrooms=classrooms)

@app.route('/classroom/<int:classroom_id>')
def classroom_detail(classroom_id):
    from datetime import datetime
    current_date = get_brazil_time().date()
    
    classroom = Classroom.query.get_or_404(classroom_id)
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).filter(
        db.or_(
            Schedule.end_date.is_(None),  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    # Get incidents for this classroom using raw SQL to avoid SQLAlchemy column issues
    incidents = []
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # First ensure the column exists
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
            
            # Use safe SQL query
            result = conn.execute(text("""
                SELECT id, classroom_id, reporter_name, reporter_email, description, 
                       created_at, is_active, is_resolved, admin_response, response_date,
                       COALESCE(hidden_from_classroom, FALSE) as hidden_from_classroom
                FROM incident 
                WHERE classroom_id = :classroom_id 
                  AND is_active = TRUE 
                  AND COALESCE(hidden_from_classroom, FALSE) = FALSE
                ORDER BY created_at DESC
            """), {'classroom_id': classroom_id})
            
            incident_data = result.fetchall()
            
            # Convert to Incident-like objects for template compatibility
            class IncidentProxy:
                def __init__(self, row):
                    self.id = row[0]
                    self.classroom_id = row[1]
                    self.reporter_name = row[2]
                    self.reporter_email = row[3]
                    self.description = row[4]
                    self.created_at = row[5]
                    self.is_active = row[6]
                    self.is_resolved = row[7]
                    self.admin_response = row[8]
                    self.response_date = row[9]
                    self.hidden_from_classroom = row[10]
            
            incidents = [IncidentProxy(row) for row in incident_data]
            
    except Exception as e:
        import logging
        logging.error(f"Incident query error: {e}")
        incidents = []
    
    return render_template('classroom.html', classroom=classroom, schedules=schedules, incidents=incidents)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['admin_authenticated'] = True
            session.permanent = True
            app.permanent_session_lifetime = timedelta(hours=2)
            flash('Login realizado com sucesso!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Senha incorreta!', 'error')
    
    return render_template('auth.html')

@app.route('/logout')
def logout():
    session.pop('admin_authenticated', None)
    flash('Logout realizado com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/install')
def install_instructions():
    """Página com instruções para instalar o aplicativo em diferentes dispositivos"""
    return render_template('install_instructions.html')

@app.route('/static/sw.js')
def service_worker():
    """Serve the service worker with correct MIME type"""
    return send_file('static/sw.js', mimetype='application/javascript')

# Error handlers para prevenir crashes
@app.errorhandler(404)
def not_found_error(error):
    """Página personalizada para erro 404"""
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Página personalizada para erro 500"""
    db.session.rollback()
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    """Página personalizada para erro 403"""
    return render_template('errors/403.html'), 403

@app.route('/edit_classroom/<int:classroom_id>', methods=['GET', 'POST'])
@require_admin_auth
def edit_classroom(classroom_id):
    current_date = get_brazil_time().date()
    
    classroom = Classroom.query.get_or_404(classroom_id)
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).filter(
        db.or_(
            Schedule.end_date.is_(None),  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    import logging
    logging.debug(f"Edit classroom showing {len(schedules)} active/current schedules for classroom {classroom_id} (expired courses hidden)")
    
    if request.method == 'POST':
        try:
            classroom.name = request.form.get('name', '')
            classroom.capacity = int(request.form.get('capacity', 0))
            classroom.has_computers = 'has_computers' in request.form
            classroom.software = request.form.get('software', '')
            classroom.description = request.form.get('description', '')

            classroom.block = request.form.get('block', '')
            classroom.admin_password = request.form.get('admin_password', '')
            
            # Handle image upload with PostgreSQL storage
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Store file data in database
                    classroom.image_data = file.read()
                    classroom.image_mimetype = file.mimetype
                    classroom.image_filename = filename
            
            # Handle Excel file upload with PostgreSQL storage
            if 'excel_file' in request.files:
                excel_file = request.files['excel_file']
                if excel_file and excel_file.filename and excel_file.filename != '' and allowed_excel_file(excel_file.filename):
                    filename = secure_filename(excel_file.filename)
                    # Store file data in database
                    classroom.excel_data = excel_file.read()
                    classroom.excel_mimetype = excel_file.mimetype
                    classroom.excel_filename = filename
                    
            classroom.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Sala atualizada com sucesso!', 'success')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar sala: {str(e)}', 'error')
            return render_template('edit_classroom.html', classroom=classroom, schedules=schedules)
    
    return render_template('edit_classroom.html', classroom=classroom, schedules=schedules)


@app.route('/download_excel/<int:classroom_id>')
def download_excel(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not classroom.excel_data:
            flash('Nenhum arquivo Excel disponível para esta sala.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        safe_filename = f"{classroom.name.replace(' ', '_')}_patrimonio.xlsx"
        return send_file(
            io.BytesIO(classroom.excel_data),
            mimetype=classroom.excel_mimetype or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        flash(f'Erro ao baixar arquivo: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/image/<int:classroom_id>')
def serve_image(classroom_id):
    """Serve images from PostgreSQL database"""
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not classroom.image_data:
            # Return default image or 404
            from flask import abort
            abort(404)
        
        return send_file(
            io.BytesIO(classroom.image_data),
            mimetype=classroom.image_mimetype or 'image/jpeg'
        )
    except Exception as e:
        from flask import abort
        abort(404)

@app.route('/upload_excel/<int:classroom_id>', methods=['POST'])
@require_admin_auth
def upload_excel(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    
    if 'excel_file' not in request.files:
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))
    
    excel_file = request.files['excel_file']
    
    if excel_file.filename == '':
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))
    
    if excel_file and excel_file.filename and allowed_excel_file(excel_file.filename):
        try:
            filename = secure_filename(excel_file.filename or '')
            
            # Store file data in database
            classroom.excel_data = excel_file.read()
            classroom.excel_mimetype = excel_file.mimetype
            classroom.excel_filename = filename
            classroom.updated_at = datetime.utcnow()
            db.session.commit()
            
            flash('Arquivo Excel carregado com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao carregar arquivo: {str(e)}', 'error')
    else:
        flash('Formato de arquivo não permitido. Use apenas arquivos .xlsx ou .xls', 'error')
    
    return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/delete_schedule/<int:schedule_id>', methods=['POST'])
@require_admin_auth
def delete_schedule(schedule_id):
    schedule = Schedule.query.get_or_404(schedule_id)
    classroom_id = schedule.classroom_id
    
    try:
        db.session.delete(schedule)
        db.session.commit()
        flash('Horário removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover horário: {str(e)}', 'error')
    
    return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/add_incident/<int:classroom_id>', methods=['POST'])
def add_incident(classroom_id):
    classroom = Classroom.query.get_or_404(classroom_id)
    
    try:
        reporter_name = request.form.get('reporter_name', '').strip()
        reporter_email = request.form.get('reporter_email', '').strip()
        description = request.form.get('description', '').strip()
        
        if not reporter_name or not reporter_email or not description:
            flash('Todos os campos são obrigatórios para registrar uma ocorrência.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        # Simple incident creation using SQLAlchemy model
        brazil_time = get_brazil_time().replace(tzinfo=None)
        
        incident = Incident(
            classroom_id=classroom_id,
            reporter_name=reporter_name,
            reporter_email=reporter_email,
            description=description
        )
        incident.created_at = brazil_time
        db.session.add(incident)
        db.session.commit()
        flash('Ocorrência registrada com sucesso! A administração será notificada.', 'success')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
    except Exception as e:
        db.session.rollback()
        import logging
        logging.error(f'Erro ao registrar ocorrência: {str(e)}')
        flash('Erro ao registrar ocorrência. Tente novamente.', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/hide_incident_from_classroom/<int:incident_id>', methods=['POST'])
@require_admin_auth
def hide_incident_from_classroom(incident_id):
    incident = Incident.query.get_or_404(incident_id)
    classroom_id = incident.classroom_id
    
    try:
        # Check if hidden_from_classroom column exists
        from sqlalchemy import text
        column_exists = False
        try:
            with db.engine.connect() as conn:
                if 'postgresql' in str(db.engine.url) or 'postgres' in str(db.engine.url):
                    result = conn.execute(text("""
                        SELECT column_name FROM information_schema.columns 
                        WHERE table_name='incident' AND column_name='hidden_from_classroom'
                    """))
                    column_exists = result.fetchone() is not None
                else:
                    try:
                        conn.execute(text("SELECT hidden_from_classroom FROM incident LIMIT 1"))
                        column_exists = True
                    except:
                        column_exists = False
        except:
            column_exists = False
        
        # Hide incident using raw SQL with column creation if needed
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            # Use raw SQL to avoid SQLAlchemy issues
            conn.execute(text("""
                UPDATE incident 
                SET hidden_from_classroom = true 
                WHERE id = :incident_id
            """), {'incident_id': incident_id})
            conn.commit()
            
        flash('Ocorrência removida da visualização da sala!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao ocultar ocorrência: {str(e)}', 'error')
    
    return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/delete_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def delete_incident(incident_id):
    incident = Incident.query.get_or_404(incident_id)
    
    # Check where we're coming from to redirect properly
    referrer = request.form.get('referrer', 'incidents_management')
    
    try:
        db.session.delete(incident)
        db.session.commit()
        flash('Ocorrência excluída permanentemente!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir ocorrência: {str(e)}', 'error')
    
    if referrer == 'classroom':
        return redirect(url_for('classroom_detail', classroom_id=incident.classroom_id))
    else:
        return redirect(url_for('incidents_management'))

@app.route('/admin/migrate_db')
@require_admin_auth  
def migrate_database():
    """Rota para migrar banco de dados - adicionar colunas faltantes"""
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            # Adicionar hidden_from_classroom
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
                flash('✅ Migração concluída com sucesso! Coluna hidden_from_classroom adicionada.', 'success')
            except Exception as e:
                if 'already exists' in str(e) or 'duplicate' in str(e):
                    flash('✅ Coluna hidden_from_classroom já existe.', 'info')  
                else:
                    flash(f'❌ Erro na migração: {str(e)}', 'error')
                    
    except Exception as e:
        flash(f'❌ Erro na migração: {str(e)}', 'error')
        
    return redirect(url_for('incidents_management'))

@app.route('/incidents_management')
@require_admin_auth
def incidents_management():
    """Admin panel for managing all incidents with filters"""
    try:
        # Get filter parameters
        status_filter = request.args.get('status', '')
        reporter_filter = request.args.get('reporter', '')
        classroom_filter = request.args.get('classroom', '')
        
        # Use raw SQL to avoid SQLAlchemy model issues with missing columns
        from sqlalchemy import text
        
        # Use simplified SQL with COALESCE to handle missing column gracefully
        base_sql = """
            SELECT id, classroom_id, reporter_name, reporter_email, description, 
                   created_at, is_active, is_resolved, admin_response, response_date,
                   COALESCE(hidden_from_classroom, FALSE) as hidden_from_classroom
        """
        where_clause = "WHERE is_active = true AND COALESCE(hidden_from_classroom, FALSE) = FALSE"
        
        # Add filters to WHERE clause
        filter_conditions = []
        params = {}
        
        if status_filter == 'pending':
            filter_conditions.append("is_resolved = false")
        elif status_filter == 'resolved':
            filter_conditions.append("is_resolved = true")
        
        if reporter_filter:
            filter_conditions.append("LOWER(reporter_name) LIKE LOWER(:reporter_filter)")
            params['reporter_filter'] = f'%{reporter_filter}%'
        
        if classroom_filter:
            try:
                classroom_id = int(classroom_filter)
                filter_conditions.append("classroom_id = :classroom_id")
                params['classroom_id'] = classroom_id
            except ValueError:
                pass
        
        # Complete SQL query
        if filter_conditions:
            full_sql = f"{base_sql} FROM incident {where_clause} AND {' AND '.join(filter_conditions)} ORDER BY created_at DESC"
        else:
            full_sql = f"{base_sql} FROM incident {where_clause} ORDER BY created_at DESC"
        
        # Execute query and convert to incident objects
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            result = conn.execute(text(full_sql), params)
            incident_rows = result.fetchall()
        
        # Convert to Incident-like objects for template compatibility
        class IncidentProxy:
            def __init__(self, row):
                self.id = row[0]
                self.classroom_id = row[1]
                self.reporter_name = row[2]
                self.reporter_email = row[3]
                self.description = row[4]
                self.created_at = row[5]
                self.is_active = row[6]
                self.is_resolved = row[7]
                self.admin_response = row[8]
                self.response_date = row[9]
                if len(row) > 10:
                    self.hidden_from_classroom = row[10]
                else:
                    self.hidden_from_classroom = False
                
                # Add classroom relationship
                self.classroom = Classroom.query.get(self.classroom_id)
        
        incidents = [IncidentProxy(row) for row in incident_rows]
        
        # Get counts using safe SQL with COALESCE
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            pending_result = conn.execute(text("""
                SELECT COUNT(*) FROM incident 
                WHERE is_active = true AND is_resolved = false 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            resolved_result = conn.execute(text("""
                SELECT COUNT(*) FROM incident 
                WHERE is_active = true AND is_resolved = true 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            
            pending_count = pending_result.scalar()
            resolved_count = resolved_result.scalar()
        
        # Get classrooms and reporters using safe queries
        classrooms = Classroom.query.all()
        
        with db.engine.connect() as conn:
            reporters_result = conn.execute(text("""
                SELECT DISTINCT reporter_name FROM incident 
                WHERE is_active = true 
                AND COALESCE(hidden_from_classroom, FALSE) = FALSE
            """))
            reporters = [row[0] for row in reporters_result.fetchall()]
        
        return render_template('incidents_management.html', 
                             incidents=incidents, 
                             pending_count=pending_count, 
                             resolved_count=resolved_count,
                             classrooms=classrooms,
                             reporters=reporters,
                             current_filters={
                                 'status': status_filter,
                                 'reporter': reporter_filter,
                                 'classroom': classroom_filter
                             })
    except Exception as e:
        import logging
        logging.error(f'Erro na gestão de ocorrências: {str(e)}')
        flash(f'Erro ao carregar gestão de ocorrências: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/respond_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def respond_incident(incident_id):
    """Admin response to an incident"""
    incident = Incident.query.get_or_404(incident_id)
    
    try:
        admin_response = request.form.get('admin_response', '').strip()
        mark_resolved = request.form.get('mark_resolved') == '1'
        
        if not admin_response:
            flash('A resposta não pode estar vazia.', 'error')
            return redirect(url_for('incidents_management'))
        
        incident.admin_response = admin_response
        incident.response_date = get_brazil_time().replace(tzinfo=None)
        
        if mark_resolved:
            incident.is_resolved = True
        
        db.session.commit()
        
        status_msg = "e marcada como resolvida" if mark_resolved else ""
        flash(f'Resposta enviada com sucesso {status_msg}!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao enviar resposta: {str(e)}', 'error')
    
    return redirect(url_for('incidents_management'))

@app.route('/resolve_incident/<int:incident_id>', methods=['POST'])
@require_admin_auth
def resolve_incident(incident_id):
    """Mark an incident as resolved"""
    incident = Incident.query.get_or_404(incident_id)
    
    try:
        incident.is_resolved = True
        incident.response_date = get_brazil_time().replace(tzinfo=None)
        db.session.commit()
        flash('Ocorrência marcada como resolvida!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao resolver ocorrência: {str(e)}', 'error')
    
    return redirect(url_for('incidents_management'))

@app.route('/incidents_pdf_report')
@require_admin_auth
def incidents_pdf_report():
    """Generate PDF report of incidents with filters"""
    try:
        # Get the same filters as incidents_management
        status_filter = request.args.get('status', '')
        reporter_filter = request.args.get('reporter', '')
        classroom_filter = request.args.get('classroom', '')
        
        # Use raw SQL for PDF report to avoid column issues
        from sqlalchemy import text
        
        # Use simplified SQL with COALESCE for PDF report
        where_clause = "WHERE is_active = true AND COALESCE(hidden_from_classroom, FALSE) = FALSE"
        
        # Add filters
        filter_conditions = []
        params = {}
        
        if status_filter == 'pending':
            filter_conditions.append("is_resolved = false")
        elif status_filter == 'resolved':
            filter_conditions.append("is_resolved = true")
        
        if reporter_filter:
            filter_conditions.append("LOWER(reporter_name) LIKE LOWER(:reporter_filter)")
            params['reporter_filter'] = f'%{reporter_filter}%'
        
        if classroom_filter:
            try:
                classroom_id = int(classroom_filter)
                filter_conditions.append("classroom_id = :classroom_id")
                params['classroom_id'] = classroom_id
            except ValueError:
                pass
        
        # Complete SQL
        if filter_conditions:
            full_sql = f"SELECT id FROM incident {where_clause} AND {' AND '.join(filter_conditions)} ORDER BY created_at DESC"
        else:
            full_sql = f"SELECT id FROM incident {where_clause} ORDER BY created_at DESC"
        
        # Get incident data directly to avoid SQLAlchemy issues
        with db.engine.connect() as conn:
            # Ensure column exists first
            try:
                conn.execute(text("ALTER TABLE incident ADD COLUMN IF NOT EXISTS hidden_from_classroom BOOLEAN DEFAULT FALSE"))
                conn.commit()
            except:
                pass
                
            # Get full incident data
            full_data_sql = f"""
                SELECT id, classroom_id, reporter_name, reporter_email, description, 
                       created_at, is_active, is_resolved, admin_response, response_date
                FROM incident {where_clause}
            """
            if filter_conditions:
                full_data_sql += f" AND {' AND '.join(filter_conditions)}"
            full_data_sql += " ORDER BY created_at DESC"
            
            result = conn.execute(text(full_data_sql), params)
            incident_data = result.fetchall()
        
        # Convert to incident-like objects
        class IncidentProxy:
            def __init__(self, row):
                self.id = row[0]
                self.classroom_id = row[1]
                self.reporter_name = row[2]
                self.reporter_email = row[3]
                self.description = row[4]
                self.created_at = row[5]
                self.is_active = row[6]
                self.is_resolved = row[7]
                self.admin_response = row[8]
                self.response_date = row[9]
                self.classroom = Classroom.query.get(self.classroom_id)
        
        incidents = [IncidentProxy(row) for row in incident_data]
        
        # Generate PDF using ReportLab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Create BytesIO buffer
        buffer = io.BytesIO()
        
        # Create the PDF object
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, textColor=colors.HexColor('#1f2937'))
        subtitle_style = ParagraphStyle('CustomSubtitle', parent=styles['Heading2'], fontSize=12, spaceAfter=20, textColor=colors.HexColor('#374151'))
        normal_style = styles['Normal']
        
        # Add title
        title = Paragraph("Relatório de Ocorrências - SENAI Morvan Figueiredo", title_style)
        elements.append(title)
        
        # Add generation date
        generation_date = f"Gerado em: {get_brazil_time().strftime('%d/%m/%Y às %H:%M')}"
        date_para = Paragraph(generation_date, normal_style)
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Add filter info if any
        filter_info = []
        if status_filter:
            filter_info.append(f"Status: {'Pendentes' if status_filter == 'pending' else 'Resolvidas'}")
        if reporter_filter:
            filter_info.append(f"Reportado por: {reporter_filter}")
        if classroom_filter:
            classroom = Classroom.query.get(int(classroom_filter))
            if classroom:
                filter_info.append(f"Sala: {classroom.name}")
        
        if filter_info:
            filter_text = "Filtros aplicados: " + ", ".join(filter_info)
            filter_para = Paragraph(filter_text, subtitle_style)
            elements.append(filter_para)
            elements.append(Spacer(1, 12))
        
        if incidents:
            # Create table data
            data = [['ID', 'Sala', 'Reportado por', 'Data', 'Status', 'Descrição']]
            
            for incident in incidents:
                status = 'Resolvida' if incident.is_resolved else 'Pendente'
                # Truncate description for table
                description = incident.description[:50] + '...' if len(incident.description) > 50 else incident.description
                data.append([
                    f"#{incident.id}",
                    incident.classroom.name,
                    incident.reporter_name,
                    incident.created_at.strftime('%d/%m/%Y') if incident.created_at else '',
                    status,
                    description
                ])
            
            # Create table
            table = Table(data, colWidths=[0.8*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 2.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            
            # Add detailed incidents
            detailed_title = Paragraph("Detalhes das Ocorrências", subtitle_style)
            elements.append(detailed_title)
            
            for incident in incidents:
                # Incident header
                incident_header = f"Ocorrência #{incident.id} - {incident.classroom.name}"
                header_para = Paragraph(incident_header, ParagraphStyle('IncidentHeader', parent=styles['Heading3'], fontSize=11, textColor=colors.HexColor('#1f2937')))
                elements.append(header_para)
                
                # Incident details
                details = f"""<b>Reportado por:</b> {incident.reporter_name} ({incident.reporter_email})<br/>
                <b>Data:</b> {incident.created_at.strftime('%d/%m/%Y às %H:%M') if incident.created_at else 'Não informada'}<br/>
                <b>Status:</b> {'Resolvida' if incident.is_resolved else 'Pendente'}<br/>
                <b>Descrição:</b> {incident.description}<br/>"""
                
                if incident.admin_response:
                    details += f"<b>Resposta do Admin:</b> {incident.admin_response}<br/>"
                    if incident.response_date:
                        details += f"<b>Data da Resposta:</b> {incident.response_date.strftime('%d/%m/%Y às %H:%M')}<br/>"
                
                details_para = Paragraph(details, normal_style)
                elements.append(details_para)
                elements.append(Spacer(1, 12))
        else:
            no_incidents = Paragraph("Nenhuma ocorrência encontrada com os filtros aplicados.", normal_style)
            elements.append(no_incidents)
        
        # Add summary
        total_incidents = len(incidents)
        pending_incidents = len([i for i in incidents if not i.is_resolved])
        resolved_incidents = len([i for i in incidents if i.is_resolved])
        
        summary = f"""<b>Resumo:</b><br/>
        Total de ocorrências: {total_incidents}<br/>
        Pendentes: {pending_incidents}<br/>
        Resolvidas: {resolved_incidents}"""
        
        summary_para = Paragraph(summary, subtitle_style)
        elements.append(Spacer(1, 20))
        elements.append(summary_para)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and create response
        pdf_data = buffer.getvalue()
        buffer.close()
        
        timestamp = get_brazil_time().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_ocorrencias_{timestamp}.pdf'
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        flash(f'Erro ao gerar relatório PDF: {str(e)}', 'error')
        return redirect(url_for('incidents_management'))


@app.route('/migrate_uploads_to_db')
@require_admin_auth
def migrate_uploads_to_db():
    """Migrate any remaining files from uploads folder to PostgreSQL database"""
    try:
        import os
        uploads_folder = 'static/uploads'
        migrated_count = 0
        
        if not os.path.exists(uploads_folder):
            flash('Pasta uploads não encontrada - todos os arquivos já estão no banco.', 'info')
            return redirect(url_for('dashboard'))
        
        # Get all classrooms that might have old file references
        classrooms = Classroom.query.all()
        
        for classroom in classrooms:
            # Check if classroom has image_filename but no image_data
            if classroom.image_filename and not classroom.image_data:
                old_image_path = os.path.join(uploads_folder, classroom.image_filename)
                if os.path.exists(old_image_path):
                    try:
                        with open(old_image_path, 'rb') as f:
                            classroom.image_data = f.read()
                            # Determine mimetype from extension
                            ext = classroom.image_filename.lower().split('.')[-1]
                            mime_map = {
                                'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                                'png': 'image/png', 'gif': 'image/gif'
                            }
                            classroom.image_mimetype = mime_map.get(ext, 'image/jpeg')
                            migrated_count += 1
                    except Exception as e:
                        import logging
                        logging.error(f"Erro ao migrar imagem {classroom.image_filename}: {e}")
            
            # Check if classroom has excel_filename but no excel_data
            # Also check Excel files with any uploads pattern
            if classroom.excel_filename and not classroom.excel_data:
                old_excel_path = os.path.join(uploads_folder, classroom.excel_filename)
                if os.path.exists(old_excel_path):
                    try:
                        with open(old_excel_path, 'rb') as f:
                            classroom.excel_data = f.read()
                            classroom.excel_mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            migrated_count += 1
                    except Exception as e:
                        import logging
                        logging.error(f"Erro ao migrar Excel {classroom.excel_filename}: {e}")
        
        db.session.commit()
        flash(f'Migração concluída! {migrated_count} arquivos movidos para o banco PostgreSQL.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro durante migração: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/add_classroom', methods=['GET', 'POST'])
@require_admin_auth
def add_classroom():
    if request.method == 'POST':
        try:
            # Handle image upload with PostgreSQL storage
            image_data = None
            image_mimetype = None
            image_filename = ''
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    image_data = file.read()
                    image_mimetype = file.mimetype
                    image_filename = filename
            
            classroom = Classroom(
                name=request.form.get('name', ''),
                capacity=int(request.form.get('capacity', 0)),
                has_computers='has_computers' in request.form,
                software=request.form.get('software', ''),
                description=request.form.get('description', ''),
                block=request.form.get('block', ''),
                image_filename=image_filename,
                admin_password=request.form.get('admin_password', '')
            )
            
            # Set image data after creation
            if image_data:
                classroom.image_data = image_data
                classroom.image_mimetype = image_mimetype
            
            db.session.add(classroom)
            db.session.commit()
            
            # Create initial schedules if provided
            initial_shift = request.form.get('initial_shift')
            if initial_shift and request.form.get('initial_course'):
                initial_days = request.form.getlist('initial_days')
                if initial_days:
                    # Process date fields
                    initial_start_date = None
                    initial_end_date = None
                    
                    start_date_value = request.form.get('initial_start_date')
                    if start_date_value and start_date_value.strip():
                        try:
                            from datetime import datetime
                            initial_start_date = datetime.strptime(start_date_value.strip(), '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    
                    end_date_value = request.form.get('initial_end_date')
                    if end_date_value and end_date_value.strip():
                        try:
                            from datetime import datetime
                            initial_end_date = datetime.strptime(end_date_value.strip(), '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    
                    for day in initial_days:
                        schedule = Schedule(
                            classroom_id=classroom.id,
                            day_of_week=int(day),
                            shift=initial_shift,
                            course_name=request.form.get('initial_course', ''),
                            instructor=request.form.get('initial_instructor', ''),
                            start_time=request.form.get('initial_start_time', ''),
                            end_time=request.form.get('initial_end_time', ''),
                            start_date=initial_start_date,
                            end_date=initial_end_date
                        )
                        db.session.add(schedule)
                    db.session.commit()
                    
                    # Enhanced success message with date info
                    date_info = ""
                    if initial_start_date and initial_end_date:
                        date_info = f" (período: {initial_start_date.strftime('%d/%m/%Y')} a {initial_end_date.strftime('%d/%m/%Y')})"
                    elif initial_start_date:
                        date_info = f" (início: {initial_start_date.strftime('%d/%m/%Y')})"
                    
                    flash(f'Sala adicionada com {len(initial_days)} horários iniciais{date_info}!', 'success')
                else:
                    flash('Sala adicionada com sucesso!', 'success')
            else:
                flash('Sala adicionada com sucesso!', 'success')
            
            return redirect(url_for('classroom_detail', classroom_id=classroom.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao adicionar sala: {str(e)}', 'error')
            
    return render_template('edit_classroom.html', classroom=None)

@app.route('/schedule_management')
@require_admin_auth
def schedule_management():
    current_date = get_brazil_time().date()
    
    classrooms = Classroom.query.all()
    
    # Only show active schedules where courses haven't ended yet
    schedules = Schedule.query.filter_by(is_active=True).filter(
        db.or_(
            Schedule.end_date.is_(None),  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    ).all()
    
    print(f"DEBUG: Schedule management showing {len(schedules)} active/current schedules (expired courses hidden)")
    
    # Organize schedules by classroom and day
    schedule_map = {}
    for schedule in schedules:
        if schedule.classroom_id not in schedule_map:
            schedule_map[schedule.classroom_id] = {}
        if schedule.day_of_week not in schedule_map[schedule.classroom_id]:
            schedule_map[schedule.classroom_id][schedule.day_of_week] = {}
        schedule_map[schedule.classroom_id][schedule.day_of_week][schedule.shift] = schedule
    
    return render_template('schedule_management.html', 
                         classrooms=classrooms, 
                         schedules=schedules,
                         schedule_map=schedule_map)

@app.route('/add_schedule', methods=['POST'])
@require_admin_auth
def add_schedule():
    try:
        classroom_id = int(request.form.get('classroom_id') or 0)
        days = request.form.getlist('days')
        
        # Handle single day submissions from first modal
        single_day = request.form.get('day_of_week')
        if single_day is not None and single_day != '':
            days = [single_day]
        
        shift = request.form.get('shift')
        course_name = request.form.get('course_name', '')
        instructor = request.form.get('instructor', '')
        start_time = request.form.get('start_time', '')
        end_time = request.form.get('end_time', '')
        start_date_str = request.form.get('start_date', '')
        end_date_str = request.form.get('end_date', '')
        
        # Parse dates with error handling
        start_date = None
        end_date = None
        if start_date_str and start_date_str.strip():
            try:
                start_date = datetime.strptime(start_date_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                flash('Data de início inválida.', 'error')
                return redirect(url_for('schedule_management'))
        if end_date_str and end_date_str.strip():
            try:
                end_date = datetime.strptime(end_date_str.strip(), '%Y-%m-%d').date()
            except ValueError:
                flash('Data de fim inválida.', 'error')
                return redirect(url_for('schedule_management'))
        
        print(f"DEBUG: Adding schedule - classroom_id: {classroom_id}, days: {days}, shift: {shift}")
        
        created_count = 0
        existing_count = 0
        
        if not days or len(days) == 0:
            flash('Nenhum dia foi selecionado!', 'error')
            return redirect(url_for('schedule_management'))
        
        for day in days:
            day_int = int(day)
            
            # Check if schedule already exists for this slot
            existing_schedule = Schedule.query.filter_by(
                classroom_id=classroom_id,
                day_of_week=day_int,
                shift=shift,
                is_active=True
            ).first()
            
            print(f"DEBUG: Day {day_int}, existing: {existing_schedule is not None}")
            
            if not existing_schedule:
                schedule = Schedule(
                    classroom_id=classroom_id,
                    day_of_week=day_int,
                    shift=shift or '',
                    course_name=course_name,
                    instructor=instructor,
                    start_time=start_time,
                    end_time=end_time,
                    start_date=start_date,
                    end_date=end_date
                )
                db.session.add(schedule)
                created_count += 1
                print(f"DEBUG: Created schedule for day {day_int}")
            else:
                existing_count += 1
                print(f"DEBUG: Schedule already exists for day {day_int}")
        
        if created_count > 0:
            db.session.commit()
            if existing_count > 0:
                flash(f'{created_count} horários adicionados, {existing_count} já existiam!', 'success')
            else:
                flash(f'{created_count} horários adicionados com sucesso!', 'success')
        elif existing_count > 0:
            flash(f'Todos os {existing_count} horários selecionados já existem!', 'warning')
        else:
            flash('Nenhum horário foi selecionado!', 'error')
        
    except Exception as e:
        print(f"DEBUG: Error in add_schedule: {str(e)}")
        db.session.rollback()
        flash(f'Erro ao adicionar horários: {str(e)}', 'error')
    
    return redirect(url_for('schedule_management'))



@app.route('/delete_classroom/<int:classroom_id>', methods=['POST'])
@require_admin_auth
def delete_classroom(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        classroom_name = classroom.name
        
        # Delete all associated schedules and incidents
        Schedule.query.filter_by(classroom_id=classroom_id).delete()
        Incident.query.filter_by(classroom_id=classroom_id).delete()
        
        # Delete the classroom
        db.session.delete(classroom)
        db.session.commit()
        
        flash(f'Sala "{classroom_name}" excluída com sucesso!', 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir sala: {str(e)}', 'error')
        return redirect(url_for('edit_classroom', classroom_id=classroom_id))

@app.route('/dashboard')
def dashboard():
    # Get filter parameters
    block_filter = request.args.get('block', '')
    instructor_filter = request.args.get('instructor', '')
    software_filter = request.args.get('software', '')
    has_computers_filter = request.args.get('has_computers', '')
    capacity_filter = request.args.get('capacity', '')
    day_filter = request.args.get('day', '')
    shift_filter = request.args.get('shift', '')
    
    # Build classroom query with filters
    classroom_query = Classroom.query
    if block_filter:
        classroom_query = classroom_query.filter(Classroom.block.ilike(f'%{block_filter}%'))
    if software_filter:
        classroom_query = classroom_query.filter(Classroom.software.ilike(f'%{software_filter}%'))
    if has_computers_filter:
        has_computers_bool = has_computers_filter.lower() == 'true'
        classroom_query = classroom_query.filter(Classroom.has_computers == has_computers_bool)
    if capacity_filter:
        capacity_ranges = {
            'small': (0, 20),
            'medium': (21, 35),
            'large': (36, 100)
        }
        if capacity_filter in capacity_ranges:
            min_cap, max_cap = capacity_ranges[capacity_filter]
            classroom_query = classroom_query.filter(
                Classroom.capacity >= min_cap,
                Classroom.capacity <= max_cap
            )
    
    classrooms = classroom_query.all()
    
    # Build schedule query with filters - ONLY SHOW ACTIVE/CURRENT COURSES
    current_date = get_brazil_time().date()
    
    schedule_query = Schedule.query.filter_by(is_active=True)
    
    # Filter out expired courses - only show courses that haven't ended yet
    schedule_query = schedule_query.filter(
        db.or_(
            Schedule.end_date.is_(None),  # No end date specified
            Schedule.end_date >= current_date  # Course hasn't ended yet
        )
    )
    
    if day_filter:
        schedule_query = schedule_query.filter(Schedule.day_of_week == int(day_filter))
    if shift_filter:
        schedule_query = schedule_query.filter(Schedule.shift == shift_filter)
    if instructor_filter:
        schedule_query = schedule_query.filter(Schedule.instructor.ilike(f'%{instructor_filter}%'))
    
    schedules = schedule_query.all()
    print(f"DEBUG: Dashboard showing {len(schedules)} active/current schedules (expired courses hidden)")
    
    # Filter classrooms by instructor if specified
    if instructor_filter:
        classroom_ids_with_instructor = set(s.classroom_id for s in schedules)
        classroom_query = classroom_query.filter(Classroom.id.in_(classroom_ids_with_instructor))
    
    # Organize schedules by classroom and day
    schedule_map = {}
    for schedule in schedules:
        if schedule.classroom_id not in schedule_map:
            schedule_map[schedule.classroom_id] = {}
        if schedule.day_of_week not in schedule_map[schedule.classroom_id]:
            schedule_map[schedule.classroom_id][schedule.day_of_week] = {}
        schedule_map[schedule.classroom_id][schedule.day_of_week][schedule.shift] = schedule
    
    # Calculate statistics
    total_slots = len(classrooms) * 23  # 6 days * 4 shifts - 1 (no Saturday night)
    occupied_slots = len([s for s in schedules if s.classroom_id in [c.id for c in classrooms]])
    free_slots = total_slots - occupied_slots
    occupancy_rate = (occupied_slots / total_slots * 100) if total_slots > 0 else 0
    
    # Get unique filter options
    all_classrooms = Classroom.query.all()
    blocks = sorted(list(set(c.block for c in all_classrooms if c.block)))
    all_schedules = Schedule.query.filter_by(is_active=True).all()
    instructors = sorted(list(set(s.instructor for s in all_schedules if s.instructor and s.instructor.strip())))
    software_list = sorted(list(set(software.strip() for c in all_classrooms if c.software for software in c.software.split(',') if software.strip())))
    
    return render_template('dashboard.html', 
                         classrooms=classrooms, 
                         schedule_map=schedule_map,
                         free_slots=free_slots,
                         occupied_slots=occupied_slots,
                         occupancy_rate=occupancy_rate,
                         blocks=blocks,
                         instructors=instructors,
                         software_list=software_list,
                         current_filters={
                             'block': block_filter,
                             'instructor': instructor_filter,
                             'software': software_filter,
                             'has_computers': has_computers_filter,
                             'capacity': capacity_filter,
                             'day': day_filter,
                             'shift': shift_filter
                         })

@app.route('/availability')
def availability():
    return redirect(url_for('dashboard'))

def get_brazil_time():
    """Get current time in Brazil timezone (UTC-3)"""
    try:
        if PYTZ_AVAILABLE and pytz and 'pytz' in globals():
            brazil_tz = pytz.timezone('America/Sao_Paulo')
            return datetime.now(brazil_tz)
        else:
            # Fallback: subtract 3 hours from UTC to approximate Brazil time
            utc_time = datetime.utcnow()
            return utc_time - timedelta(hours=3)
    except Exception:
        # Ultimate fallback - just use UTC
        return datetime.utcnow()

def get_current_shift():
    """Get the current shift based on Brazil time"""
    now = get_brazil_time()
    current_hour = now.hour
    current_minute = now.minute
    current_time_minutes = current_hour * 60 + current_minute
    
    # Use real current time
    # current_hour = 14  # FOR TESTING only
    # current_minute = 30
    # current_time_minutes = current_hour * 60 + current_minute
    
    print(f"DEBUG: Current time: {current_hour}:{current_minute:02d} ({current_time_minutes} minutes)")
    
    # Define shift time ranges in minutes
    # Morning: 7:30-12:00 (450-720 minutes)
    # Afternoon: 13:00-18:00 (780-1080 minutes)  
    # Night: 18:30-22:30 (1110-1350 minutes)
    # Fullday: 8:00-17:00 (480-1020 minutes) - as per schedule data
    
    current_shifts = []
    
    if 450 <= current_time_minutes <= 720:  # Morning
        current_shifts.append('morning')
        print(f"DEBUG: Added morning shift")
    if 780 <= current_time_minutes <= 1080:  # Afternoon
        current_shifts.append('afternoon')
        print(f"DEBUG: Added afternoon shift")
    if 1110 <= current_time_minutes <= 1350:  # Night
        current_shifts.append('night')
        print(f"DEBUG: Added night shift")
    # Fullday: check if within fullday hours (8:00-17:00)
    if 480 <= current_time_minutes <= 1020:  # Fullday
        current_shifts.append('fullday')
        print(f"DEBUG: Added fullday shift")
        
    print(f"DEBUG: Current active shifts: {current_shifts}")
    return current_shifts

def get_availability_for_date(target_date=None, shift_filter=None):
    """Helper function to get room availability for a specific date and optional shift"""
    if target_date is None:
        target_date = get_brazil_time()
    
    # Get day of week (0=Monday, 6=Sunday)
    target_day = target_date.weekday()
    target_date_only = target_date.date()
    
    classrooms = Classroom.query.all()
    
    print(f"DEBUG: Checking availability for date: {target_date_only}, day of week: {target_day}")
    
    # Check if it's Sunday
    if target_day == 6:  # Sunday
        return {
            'available_rooms': classrooms,
            'occupied_rooms': [],
            'period_description': "Domingo - Escola fechada",
            'total_rooms': len(classrooms)
        }
    
    # If no shift filter is provided and we're checking current time, get current shifts
    if shift_filter is None or shift_filter == 'all':
        # If checking current date, determine the primary current shift
        if target_date.date() == get_brazil_time().date():
            current_shifts = get_current_shift()
            print(f"DEBUG: Checking current date, active shifts: {current_shifts}")
            
            if not current_shifts:  # Outside operating hours
                return {
                    'available_rooms': classrooms,
                    'occupied_rooms': [],
                    'period_description': "Fora do horário de funcionamento",
                    'total_rooms': len(classrooms)
                }
            
            # Determine which shift to check based on current time
            # Priority: specific shifts (morning, afternoon, night) over fullday
            primary_shift = None
            
            # Check for specific time-based shifts first
            if 'morning' in current_shifts:
                primary_shift = 'morning'
            elif 'afternoon' in current_shifts:
                primary_shift = 'afternoon'
            elif 'night' in current_shifts:
                primary_shift = 'night'
            elif 'fullday' in current_shifts:
                primary_shift = 'fullday'
            
            occupied_schedules = []
            
            if primary_shift:
                # Get schedules for the primary shift only - PRECISE DATE CHECKING
                all_schedules = Schedule.query.filter_by(
                    day_of_week=target_day,
                    shift=primary_shift,
                    is_active=True
                ).all()
                
                # Filter by actual course dates
                active_schedules = []
                for schedule in all_schedules:
                    if schedule.start_date and schedule.end_date:
                        if schedule.start_date <= target_date_only <= schedule.end_date:
                            active_schedules.append(schedule)
                            print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) is ACTIVE (course runs {schedule.start_date} to {schedule.end_date})")
                        else:
                            print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) is EXPIRED/FUTURE (course runs {schedule.start_date} to {schedule.end_date}, today is {target_date_only})")
                    else:
                        # If no dates specified, consider it active (backward compatibility)
                        active_schedules.append(schedule)
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift}) has no date restrictions, treating as active")
                
                occupied_schedules.extend(active_schedules)
                print(f"DEBUG: Using primary shift '{primary_shift}', found {len(active_schedules)} ACTIVE schedules out of {len(all_schedules)} total")
                
                # CRITICAL LOGIC: Only add fullday schedules if we're checking for CURRENT time
                # This prevents fullday classes from appearing when user filters by specific shift
                if primary_shift in ['morning', 'afternoon'] and target_date.date() == get_brazil_time().date():
                    all_fullday_schedules = Schedule.query.filter_by(
                        day_of_week=target_day,
                        shift='fullday',
                        is_active=True
                    ).all()
                    
                    active_fullday_schedules = []
                    for schedule in all_fullday_schedules:
                        if schedule.start_date and schedule.end_date:
                            if schedule.start_date <= target_date_only <= schedule.end_date:
                                active_fullday_schedules.append(schedule)
                                print(f"DEBUG: Fullday schedule {schedule.id} is ACTIVE (overlaps with current {primary_shift} shift)")
                            else:
                                print(f"DEBUG: Fullday schedule {schedule.id} is EXPIRED/FUTURE")
                        else:
                            active_fullday_schedules.append(schedule)
                    
                    occupied_schedules.extend(active_fullday_schedules)
                    print(f"DEBUG: Added {len(active_fullday_schedules)} ACTIVE fullday schedules for CURRENT TIME overlap")
                else:
                    print(f"DEBUG: Skipping fullday overlap - not checking current time or not morning/afternoon shift")
        else:
            # For other dates (future/past), check ALL shifts to get complete availability picture
            print(f"DEBUG: Checking NON-CURRENT date {target_date_only} - checking ALL shifts for complete availability")
            all_schedules = Schedule.query.filter_by(day_of_week=target_day, is_active=True).all()
            
            active_schedules = []
            for schedule in all_schedules:
                if schedule.start_date and schedule.end_date:
                    if schedule.start_date <= target_date_only <= schedule.end_date:
                        active_schedules.append(schedule)
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is ACTIVE on {target_date_only}")
                    else:
                        print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is EXPIRED/FUTURE on {target_date_only}")
                else:
                    active_schedules.append(schedule)
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) has no date restrictions, treating as active")
            
            occupied_schedules = active_schedules
            print(f"DEBUG: Future/past date check - found {len(active_schedules)} ACTIVE schedules out of {len(all_schedules)} total")
    else:
        # Apply specific shift filter - ULTRA PRECISE: ONLY show rooms occupied by EXACTLY that shift
        print(f"DEBUG: PRECISE FILTER MODE - Looking for shift '{shift_filter}' on {target_date_only}")
        
        # Get schedules that EXACTLY match the requested shift
        all_schedules = Schedule.query.filter_by(
            day_of_week=target_day,
            shift=shift_filter,
            is_active=True
        ).all()
        
        active_schedules = []
        for schedule in all_schedules:
            if schedule.start_date and schedule.end_date:
                if schedule.start_date <= target_date_only <= schedule.end_date:
                    active_schedules.append(schedule)
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is ACTIVE with exact shift filter '{shift_filter}'")
                else:
                    print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) is EXPIRED/FUTURE with shift filter")
            else:
                active_schedules.append(schedule)
                print(f"DEBUG: Schedule {schedule.id} ({schedule.shift} - {schedule.course_name}) has no date restrictions, treating as active for shift filter")
        
        occupied_schedules = active_schedules
        
        # CRITICAL: For specific shift filters, we need to consider fullday classes as conflicts too
        # BUT only when the user is NOT specifically looking for fullday
        if shift_filter != 'fullday':
            print(f"DEBUG: Checking if any fullday classes conflict with '{shift_filter}' filter")
            
            # Get fullday schedules for this day
            all_fullday_schedules = Schedule.query.filter_by(
                day_of_week=target_day,
                shift='fullday',
                is_active=True
            ).all()
            
            active_fullday_schedules = []
            for schedule in all_fullday_schedules:
                if schedule.start_date and schedule.end_date:
                    if schedule.start_date <= target_date_only <= schedule.end_date:
                        active_fullday_schedules.append(schedule)
                        print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) CONFLICTS with '{shift_filter}' filter")
                    else:
                        print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) is EXPIRED/FUTURE, no conflict")
                else:
                    active_fullday_schedules.append(schedule)
                    print(f"DEBUG: Fullday schedule {schedule.id} ({schedule.course_name}) has no date restrictions, treating as conflict")
            
            occupied_schedules.extend(active_fullday_schedules)
            print(f"DEBUG: Added {len(active_fullday_schedules)} conflicting fullday schedules to '{shift_filter}' filter")
        
        print(f"DEBUG: EXACT SHIFT FILTER '{shift_filter}' - found {len(active_schedules)} exact matches + {len(occupied_schedules) - len(active_schedules)} conflicting schedules")
    
    occupied_classroom_ids = set(schedule.classroom_id for schedule in occupied_schedules)
    
    available_rooms = [room for room in classrooms if room.id not in occupied_classroom_ids]
    occupied_rooms = [room for room in classrooms if room.id in occupied_classroom_ids]
    
    # Build period description
    days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
    day_name = days[target_day]
    
    if shift_filter and shift_filter != 'all':
        shift_names = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
        period_description = f"{day_name} - {shift_names.get(shift_filter, shift_filter)} (Filtro Específico)"
    elif target_date.date() == get_brazil_time().date() and (shift_filter is None or shift_filter == 'all'):
        # Show current period
        current_shifts = get_current_shift()
        if current_shifts:
            shift_names = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
            active_shift_names = [shift_names.get(shift, shift) for shift in current_shifts]
            period_description = f"{day_name} - {', '.join(set(active_shift_names))} (Agora)"
        else:
            period_description = f"{day_name} - Fora do horário"
    else:
        period_description = f"{day_name} - Todos os turnos"
    
    return {
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'occupied_schedules': occupied_schedules,
        'period_description': period_description,
        'total_rooms': len(classrooms)
    }

@app.route('/available_now')
def available_now():
    # Get query parameters for date and shift filtering
    date_param = request.args.get('date')
    shift_param = request.args.get('shift', 'all')
    
    # Parse target date
    if date_param:
        try:
            from datetime import datetime
            target_date = datetime.strptime(date_param, '%Y-%m-%d')
        except ValueError:
            target_date = get_brazil_time()
    else:
        target_date = get_brazil_time()
    
    # Get availability data
    availability_data = get_availability_for_date(target_date, shift_param)
    
    # Format date for display
    formatted_date = target_date.strftime('%d/%m/%Y')
    
    return render_template('available_now.html', 
                         available_rooms=availability_data['available_rooms'],
                         occupied_rooms=availability_data.get('occupied_rooms', []),
                         occupied_schedules=availability_data.get('occupied_schedules', []),
                         current_period=availability_data['period_description'],
                         total_rooms=availability_data['total_rooms'],
                         selected_date=formatted_date,
                         selected_date_iso=target_date.strftime('%Y-%m-%d'),
                         selected_shift=shift_param)

@app.route('/generate_pdf/<int:classroom_id>')
def generate_pdf(classroom_id):
    if not generate_classroom_pdf:
        flash('Geração de PDF não está disponível no momento.', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
    classroom = Classroom.query.get_or_404(classroom_id)
    schedules = Schedule.query.filter_by(classroom_id=classroom_id, is_active=True).all()
    
    try:
        pdf_buffer = generate_classroom_pdf(classroom, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'sala_{classroom.name.replace(" ", "_")}.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

@app.route('/generate_general_report')
def generate_general_report_route():
    if not generate_general_report:
        flash('Geração de relatórios não está disponível no momento.', 'error')
        return redirect(url_for('dashboard'))
        
    try:
        classrooms = Classroom.query.all()
        schedules = Schedule.query.filter_by(is_active=True).all()
        
        pdf_buffer = generate_general_report(classrooms, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name='relatorio_geral.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate_availability_report')
def generate_availability_report_route():
    if not generate_availability_report:
        flash('Geração de relatórios não está disponível no momento.', 'error')
        return redirect(url_for('dashboard'))
        
    try:
        classrooms = Classroom.query.all()
        schedules = Schedule.query.filter_by(is_active=True).all()
        
        pdf_buffer = generate_availability_report(classrooms, schedules)
        
        return send_file(
            io.BytesIO(pdf_buffer.getvalue()),
            mimetype='application/pdf',
            as_attachment=True,
            download_name='relatorio_disponibilidade.pdf'
        )
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate_qr/<int:classroom_id>')
def generate_qr(classroom_id):
    try:
        classroom = Classroom.query.get_or_404(classroom_id)
        
        if not generate_qr_code:
            flash('Geração de QR code não está disponível no momento.', 'error')
            return redirect(url_for('classroom_detail', classroom_id=classroom_id))
        
        # Generate the full URL for the classroom
        classroom_url = request.url_root.rstrip('/') + url_for('classroom_detail', classroom_id=classroom_id)
        
        qr_buffer = generate_qr_code(classroom_url, classroom.name)
        safe_filename = f'qr_sala_{classroom.name.replace(" ", "_").replace("/", "_")}.png'
        
        return send_file(
            io.BytesIO(qr_buffer.getvalue()),
            mimetype='image/png',
            as_attachment=True,
            download_name=safe_filename
        )
    except Exception as e:
        flash(f'Erro ao gerar QR code: {str(e)}', 'error')
        return redirect(url_for('classroom_detail', classroom_id=classroom_id))

# Exportação para Excel - versão corrigida
@app.route('/export_excel')
def export_excel():
    try:
        # Check if openpyxl is available
        if not openpyxl:
            flash('Funcionalidade de Excel não está disponível no momento.', 'error')
            return redirect(url_for('dashboard'))
            
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        
        # Sheet 1: Classrooms
        ws1 = wb.active
        ws1.title = "Salas de Aula"
        
        # Headers for classrooms
        headers1 = ['ID', 'Nome', 'Capacidade', 'Bloco', 'Tem Computadores', 'Softwares', 'Descrição']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data for classrooms
        classrooms = Classroom.query.all()
        for row, classroom in enumerate(classrooms, 2):
            ws1.cell(row=row, column=1).value = classroom.id
            ws1.cell(row=row, column=2).value = classroom.name
            ws1.cell(row=row, column=3).value = classroom.capacity
            ws1.cell(row=row, column=4).value = classroom.block
            ws1.cell(row=row, column=5).value = 'Sim' if classroom.has_computers else 'Não'
            ws1.cell(row=row, column=6).value = classroom.software
            ws1.cell(row=row, column=7).value = classroom.description
        
        # Auto-fit columns
        for column_cells in ws1.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Schedules
        ws2 = wb.create_sheet(title="Horários")
        
        # Headers for schedules
        headers2 = ['ID', 'Sala', 'Dia da Semana', 'Turno', 'Curso', 'Professor', 'Início', 'Fim', 'Ativo']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data for schedules
        schedules = Schedule.query.all()
        days = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        shifts = {'morning': 'Manhã', 'afternoon': 'Tarde', 'fullday': 'Integral', 'night': 'Noite'}
        
        for row, schedule in enumerate(schedules, 2):
            classroom = Classroom.query.get(schedule.classroom_id)
            ws2.cell(row=row, column=1).value = schedule.id
            ws2.cell(row=row, column=2).value = classroom.name if classroom else 'N/A'
            ws2.cell(row=row, column=3).value = days[schedule.day_of_week]
            ws2.cell(row=row, column=4).value = shifts.get(schedule.shift, schedule.shift)
            ws2.cell(row=row, column=5).value = schedule.course_name
            ws2.cell(row=row, column=6).value = schedule.instructor
            ws2.cell(row=row, column=7).value = schedule.start_time
            ws2.cell(row=row, column=8).value = schedule.end_time
            ws2.cell(row=row, column=9).value = 'Sim' if schedule.is_active else 'Não'
        
        # Auto-fit columns for sheet 2
        for column_cells in ws2.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Statistics
        ws3 = wb.create_sheet(title="Estatísticas")
        ws3.cell(row=1, column=1).value = "Estatística"
        ws3.cell(row=1, column=2).value = "Valor"
        
        # Style headers
        for col in [1, 2]:
            cell = ws3.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Statistics data
        total_classrooms = len(classrooms)
        total_schedules = len([s for s in schedules if s.is_active])
        total_slots = total_classrooms * 23  # 6 days * 4 shifts - 1 (no Saturday night)
        occupancy_rate = (total_schedules / total_slots * 100) if total_slots > 0 else 0
        
        stats_data = [
            ['Total de Salas', total_classrooms],
            ['Total de Horários Ativos', total_schedules],
            ['Taxa de Ocupação (%)', f"{occupancy_rate:.1f}%"],
            ['Salas com Computadores', len([c for c in classrooms if c.has_computers])],
            ['Salas sem Computadores', len([c for c in classrooms if not c.has_computers])]
        ]
        
        for row, (stat, value) in enumerate(stats_data, 2):
            ws3.cell(row=row, column=1).value = stat
            ws3.cell(row=row, column=2).value = value
        
        # Auto-fit columns for sheet 3
        for column_cells in ws3.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_senai_{timestamp}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except Exception as e:
        flash(f'Erro ao gerar Excel: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/export_filtered_excel')
def export_filtered_excel():
    try:
        # Get the same filters as dashboard
        block_filter = request.args.get('block', '')
        # Remove floor filter as it doesn't exist in the model
        has_computers_filter = request.args.get('has_computers', '')
        capacity_filter = request.args.get('capacity', '')
        day_filter = request.args.get('day', '')
        shift_filter = request.args.get('shift', '')
        
        # Build filtered queries
        classroom_query = Classroom.query
        if block_filter:
            classroom_query = classroom_query.filter(Classroom.block == block_filter)
        # Remove floor filter as it doesn't exist in the model
        if has_computers_filter:
            has_computers_bool = has_computers_filter.lower() == 'true'
            classroom_query = classroom_query.filter(Classroom.has_computers == has_computers_bool)
        if capacity_filter:
            capacity_ranges = {
                'small': (0, 20),
                'medium': (21, 35),
                'large': (36, 100)
            }
            if capacity_filter in capacity_ranges:
                min_cap, max_cap = capacity_ranges[capacity_filter]
                classroom_query = classroom_query.filter(
                    Classroom.capacity >= min_cap,
                    Classroom.capacity <= max_cap
                )
        
        filtered_classrooms = classroom_query.all()
        
        # Build schedule query with filters
        schedule_query = Schedule.query.filter_by(is_active=True)
        if day_filter:
            schedule_query = schedule_query.filter(Schedule.day_of_week == int(day_filter))
        if shift_filter:
            schedule_query = schedule_query.filter(Schedule.shift == shift_filter)
        
        filtered_schedules = schedule_query.all()
        
        # Create Excel file similar to export_excel but with filtered data
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Salas Filtradas"
        
        # Headers
        headers = ['ID', 'Nome', 'Capacidade', 'Bloco', 'Tem Computadores', 'Softwares', 'Descrição']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, classroom in enumerate(filtered_classrooms, 2):
            ws1.cell(row=row, column=1).value = classroom.id
            ws1.cell(row=row, column=2).value = classroom.name
            ws1.cell(row=row, column=3).value = classroom.capacity
            ws1.cell(row=row, column=4).value = classroom.block
            ws1.cell(row=row, column=5).value = 'Sim' if classroom.has_computers else 'Não'
            ws1.cell(row=row, column=6).value = classroom.software
            ws1.cell(row=row, column=7).value = classroom.description
        
        # Auto-fit columns
        for column_cells in ws1.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'relatorio_filtrado_{timestamp}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
    except Exception as e:
        flash(f'Erro ao gerar Excel filtrado: {str(e)}', 'error')
        return redirect(url_for('dashboard'))